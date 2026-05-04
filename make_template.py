import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import base64, io, sys, os

# ── Colores (hex sin #) ───────────────────────────────────────────────────────
C_BG_PAGE   = "1A1A2E"   # fondo de página muy oscuro
C_HEADER    = "16213E"   # encabezado de tabla (EJERCICIOS)
C_DAY       = "0F3460"   # etiqueta lateral Día N
C_BAJA      = "2D4059"   # banda BAJA
C_MEDIA     = "1B4332"   # banda MEDIA
C_MALT      = "4A235A"   # banda MEDIA-ALTA
C_ALTA      = "641220"   # banda ALTA
C_SUBHDR    = "2C2C3E"   # sub-encabezados (MC1, S, R…)
C_EX_ODD    = "1E1E2E"   # filas ejercicio impares
C_EX_EVEN   = "252538"   # filas ejercicio pares
C_ACC_HDR   = "333344"   # fila ACCESORIOS
C_WARMUP    = "0D0D1A"   # fila calentamiento
C_TITLE     = "101020"   # filas de título superiores
C_WHITE     = "FFFFFF"
C_LIGHT     = "AAAACC"
C_MUTED     = "666688"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_WHITE, size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color="444466"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ── Workbook ─────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Rutina"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "0F3460"

# ── Anchos de columna ─────────────────────────────────────────────────────────
# A=etiqueta día, B=ejercicio, luego 4 bandas × 6 cols cada una
ws.column_dimensions["A"].width = 3.5
ws.column_dimensions["B"].width = 28

band_cols = {
    "baja":  ["C","D","E","F","G","H"],
    "media": ["I","J","K","L","M","N"],
    "malt":  ["O","P","Q","R","S","T"],
    "alta":  ["U","V","W","X","Y","Z"],
}
sub_labels = ["MC", "t'", "S", "R", "RIR", "%"]
col_widths  = [4, 3.5, 5, 5, 4.5, 6]

for band_cols_list in band_cols.values():
    for i, col in enumerate(band_cols_list):
        ws.column_dimensions[col].width = col_widths[i]

# ── Helpers ───────────────────────────────────────────────────────────────────
def apply(cell, bg=None, fnt=None, aln=None, brd=None):
    if bg:  cell.fill      = fill(bg)
    if fnt: cell.font      = fnt
    if aln: cell.alignment = aln
    if brd: cell.border    = brd

def set_row_height(row, h):
    ws.row_dimensions[row].height = h

def paint_row(row, col_start="A", col_end="Z", bg=C_BG_PAGE):
    for c in range(ord(col_start), ord(col_end)+1):
        ws[f"{chr(c)}{row}"].fill = fill(bg)

def write_band_headers(header_row, sub_row):
    band_names = {"baja":"BAJA","media":"MEDIA","malt":"MEDIA-ALTA","alta":"ALTA"}
    band_colors = {"baja":C_BAJA,"media":C_MEDIA,"malt":C_MALT,"alta":C_ALTA}
    for band, cols in band_cols.items():
        first, last = cols[0], cols[-1]
        ws.merge_cells(f"{first}{header_row}:{last}{header_row}")
        cell = ws[f"{first}{header_row}"]
        cell.value     = band_names[band]
        cell.fill      = fill(band_colors[band])
        cell.font      = font(bold=True, size=9)
        cell.alignment = align()
        cell.border    = thin_border()
        for col, lbl in zip(cols, sub_labels):
            c = ws[f"{col}{sub_row}"]
            c.value     = lbl
            c.fill      = fill(C_SUBHDR)
            c.font      = font(bold=True, color=C_LIGHT, size=8)
            c.alignment = align()
            c.border    = thin_border()

def write_ex_header(header_row, sub_row):
    # EJERCICIOS label spanning A+B
    ws.merge_cells(f"A{header_row}:B{header_row}")
    cell = ws.cell(row=header_row, column=1)
    cell.value     = "EJERCICIOS"
    cell.fill      = fill(C_HEADER)
    cell.font      = font(bold=True, size=9)
    cell.alignment = align(h="left")
    cell.border    = thin_border()
    ws[f"A{sub_row}"].fill = fill(C_SUBHDR)
    ws[f"B{sub_row}"].fill = fill(C_SUBHDR)
    ws[f"B{sub_row}"].font = font(color=C_MUTED, size=8)
    ws[f"B{sub_row}"].alignment = align(h="left")

def write_day_block(day_num, warmup_row, header_row, sub_row, ex_start, acc_start):
    # ── Calentamiento ──
    set_row_height(warmup_row, 22)
    ws.merge_cells(f"A{warmup_row}:Z{warmup_row}")
    c = ws.cell(row=warmup_row, column=1)
    c.value     = ""   # se llena desde JS
    c.fill      = fill(C_WARMUP)
    c.font      = font(bold=False, color="FFCC44", size=8.5, italic=True)
    c.alignment = align(h="left", wrap=True)
    c.border    = thin_border("222233")

    # ── Headers de banda ──
    set_row_height(header_row, 16)
    set_row_height(sub_row, 13)
    paint_row(header_row)
    paint_row(sub_row)
    write_ex_header(header_row, sub_row)
    write_band_headers(header_row, sub_row)

    # ── Etiqueta lateral "Día N" (ocupa filas ex_start..acc_start+1) ──
    merge_top = ex_start
    merge_bot = acc_start + 1
    ws.merge_cells(f"A{merge_top}:A{merge_bot}")
    day_cell = ws[f"A{merge_top}"]
    day_cell.value     = f"D\ní\na\n{day_num}"
    day_cell.fill      = fill(C_DAY)
    day_cell.font      = font(bold=True, size=9)
    day_cell.alignment = Alignment(horizontal="center", vertical="center",
                                   text_rotation=0, wrap_text=True)
    day_cell.border    = thin_border()

    # ── Filas de ejercicios ──
    for i in range(8):
        row = ex_start + i
        set_row_height(row, 14)
        bg = C_EX_ODD if i % 2 == 0 else C_EX_EVEN
        b = ws[f"B{row}"]
        b.value     = ""   # se llena desde JS
        b.fill      = fill(bg)
        b.font      = font(size=8.5)
        b.alignment = align(h="left")
        b.border    = thin_border()
        for band, cols in band_cols.items():
            for col in cols:
                c = ws[f"{col}{row}"]
                c.fill      = fill(bg)
                c.font      = font(size=8, color=C_LIGHT)
                c.alignment = align()
                c.border    = thin_border()

    # ── Fila ACCESORIOS (columna A ya está mergeada con etiqueta del día, usar B:Z) ──
    acc_label_row = acc_start - 1
    set_row_height(acc_label_row, 13)
    ws.merge_cells(f"B{acc_label_row}:Z{acc_label_row}")
    c = ws.cell(row=acc_label_row, column=2)
    c.value     = "ACCESORIOS"
    c.fill      = fill(C_ACC_HDR)
    c.font      = font(bold=True, size=8, color=C_LIGHT)
    c.alignment = align(h="left")
    c.border    = thin_border()
    ws.cell(row=acc_label_row, column=1).fill = fill(C_DAY)

    # ── Filas de accesorios ──
    for i in range(2):
        row = acc_start + i
        set_row_height(row, 13)
        b = ws[f"B{row}"]
        b.value     = ""
        b.fill      = fill(C_EX_EVEN)
        b.font      = font(size=8.5)
        b.alignment = align(h="left")
        b.border    = thin_border()
        for band, cols in band_cols.items():
            for col in cols:
                c = ws[f"{col}{row}"]
                c.fill      = fill(C_EX_EVEN)
                c.font      = font(size=8, color=C_MUTED)
                c.alignment = align()
                c.border    = thin_border()

# ── Filas de título (1 y 2) ───────────────────────────────────────────────────
for r in [1, 2]:
    set_row_height(r, 20)
    paint_row(r, bg=C_TITLE)

ws.merge_cells("A1:N1")
ws["A1"].value     = "Inicio de Rutina: "   # JS reemplaza esto
ws["A1"].fill      = fill(C_TITLE)
ws["A1"].font      = font(bold=False, size=9, color=C_LIGHT)
ws["A1"].alignment = align(h="left")

ws.merge_cells("A2:N2")
ws["A2"].value     = "Nombre:"
ws["A2"].fill      = fill(C_TITLE)
ws["A2"].font      = font(bold=True, size=10)
ws["A2"].alignment = align(h="left")

# ── Logo (O1:Z2) ─────────────────────────────────────────────────────────────
logo_path = r"D:\Descargas\gym\exodium.png"
if os.path.exists(logo_path):
    ws.merge_cells("O1:Z2")
    ws["O1"].fill = fill(C_TITLE)
    img = XLImage(logo_path)
    # Ajustar tamaño para que entre en el espacio O1:Z2 (~200x40 px aprox)
    img.width  = 160
    img.height = 40
    ws.add_image(img, "O1")
else:
    ws.merge_cells("O1:Z2")
    logo_cell = ws["O1"]
    logo_cell.value     = "EXODIUM"
    logo_cell.fill      = fill(C_TITLE)
    logo_cell.font      = font(bold=True, size=13, color="FFCC44")
    logo_cell.alignment = align(h="right")

# ── Bloques por día ───────────────────────────────────────────────────────────
# DAY_LAYOUT coincide con el JS: warmup, exStart, accStart
# El header_row = warmup+1, sub_row = warmup+2, acc_label = accStart-1
DAY_LAYOUT = [
    dict(day=1, warmup=3,  ex_start=6,  acc_start=15),
    dict(day=2, warmup=17, ex_start=20, acc_start=29),
    dict(day=3, warmup=31, ex_start=34, acc_start=43),
    dict(day=4, warmup=45, ex_start=48, acc_start=57),
    dict(day=5, warmup=59, ex_start=62, acc_start=71),
]

for d in DAY_LAYOUT:
    write_day_block(
        day_num    = d["day"],
        warmup_row = d["warmup"],
        header_row = d["warmup"] + 1,
        sub_row    = d["warmup"] + 2,
        ex_start   = d["ex_start"],
        acc_start  = d["acc_start"],
    )

# ── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "C3"

# ── Guardar y convertir a base64 ─────────────────────────────────────────────
out_path = r"D:\Descargas\gym\template_new_3.xlsx"
wb.save(out_path)
print(f"Template guardado en {out_path}")

buf = io.BytesIO()
wb.save(buf)
b64 = base64.b64encode(buf.getvalue()).decode()
print(f"\nBase64 ({len(b64)} chars):\n{b64[:80]}...")

# Guardar base64 en archivo para copiar fácil
with open(r"D:\Descargas\gym\template_b64.txt", "w") as f:
    f.write(b64)
print(f"\nBase64 guardado en template_b64.txt")
